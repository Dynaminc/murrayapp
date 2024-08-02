from django.contrib import admin # noqa
from .models import *
from django.http import HttpResponse
from openpyxl import Workbook


# Register your models here.
admin.site.register(Company)
admin.site.register(Missing)
admin.site.register(MiniCombination)
admin.site.register(Nonday)
admin.site.register(Earning)
admin.site.register(Cronny)

from .models import Combination

class SymbolFilter(admin.SimpleListFilter):
    title = 'Symbol'
    parameter_name = 'symbol'

    def lookups(self, request, model_admin):
        symbols = Combination.objects.values_list('symbol', flat=True).distinct()
        return [(symbol, symbol) for symbol in symbols]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(symbol=self.value())
        return queryset

class DateTimeFilter(admin.SimpleListFilter):
    title = 'Date Time'
    parameter_name = 'date_time'

    def lookups(self, request, model_admin):
        date_times = Combination.objects.values_list('date_time', flat=True).distinct()
        return [(date_time, date_time) for date_time in date_times]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(date_time=self.value())
        return queryset

class AscendingDescendingStdevFilter(admin.SimpleListFilter):
    title = 'Stdev Ascending/Descending'
    parameter_name = 'stdev_asc_desc'

    def lookups(self, request, model_admin):
        return [
            ('asc', 'Ascending'),
            ('desc', 'Descending'),
        ]

    def queryset(self, request, queryset):
        if self.value() == 'asc':
            return queryset.order_by('avg')
        elif self.value() == 'desc':
            return queryset.order_by('-avg')
        return queryset

class AscendingDescendingZScoreFilter(admin.SimpleListFilter):
    title = 'Z-Score Ascending/Descending'
    parameter_name = 'z_score_asc_desc'

    def lookups(self, request, model_admin):
        return [
            ('asc', 'Ascending'),
            ('desc', 'Descending'),
        ]

    def queryset(self, request, queryset):
        if self.value() == 'asc':
            return queryset.order_by('avg')
        elif self.value() == 'desc':
            return queryset.order_by('-avg')
        return queryset

# class CombinationAdmin(admin.ModelAdmin):
#     list_display = ('symbol', 'strike', 'avg', 'stdev', 'z_score', 'date_time')
#     list_filter = (SymbolFilter, DateTimeFilter, AscendingDescendingStdevFilter, AscendingDescendingZScoreFilter)
#     actions = ['export_to_excel']

#     def export_to_excel(self, request, queryset):
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=combinations.xlsx'
        
#         workbook = Workbook()
#         worksheet = workbook.active
#         worksheet.append(['Symbol', 'Strike', 'Avg', 'Stdev', 'Z-Score', 'Date Time'])

#         for obj in queryset:
#             worksheet.append([obj.symbol, obj.strike, obj.avg, obj.stdev, obj.z_score, obj.date_time])

#         workbook.save(response)
#         return response

#     export_to_excel.short_description = 'Export selected combinations to Excel'

# admin.site.register(Combination, CombinationAdmin)

class StockSymbolFilter(admin.SimpleListFilter):
    title = 'Symbol'
    parameter_name = 'symbol'

    def lookups(self, request, model_admin):
        symbols = Stock.objects.values_list('symbol', flat=True).distinct()
        return [(symbol, symbol) for symbol in symbols]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(symbol=self.value())
        return queryset

class StockDateTimeFilter(admin.SimpleListFilter):
    title = 'Date Time'
    parameter_name = 'date_time'

    def lookups(self, request, model_admin):
        date_times = Stock.objects.values_list('date_time', flat=True).distinct()
        return [(date_time, date_time) for date_time in date_times]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(date_time=self.value())
        return queryset

class CombinationDateFilter(admin.SimpleListFilter):
    title = 'Date'
    parameter_name = 'date'

    def lookups(self, request, model_admin):
        dates = Combination.objects.datetimes('date_time', 'day', order='DESC')
        return [(date.strftime('%Y-%m-%d'), date.strftime('%Y-%m-%d')) for date in dates]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(date_time__date=self.value())
        return queryset
    
class CombinationAdmin(admin.ModelAdmin):
    list_display = ('symbol', 'strike', 'avg', 'stdev', 'z_score', 'date_time')
    list_filter = (SymbolFilter, DateTimeFilter, AscendingDescendingStdevFilter, AscendingDescendingZScoreFilter, CombinationDateFilter)

    def export_to_excel(self, request, queryset):
        queryset = queryset.order_by('date_time')  # Sort by datetime
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=combinations.xlsx'
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Symbol', 'Strike', 'Avg', 'Stdev', 'Z-Score', 'Date Time'])

        for obj in queryset:
            worksheet.append([obj.symbol, obj.strike, obj.avg, obj.stdev, obj.z_score, obj.date_time])

        workbook.save(response)
        return response

    export_to_excel.short_description = 'Export selected combinations to Excel'

    actions = ['export_to_excel']

admin.site.register(Combination, CombinationAdmin)

# StockAdmin
class StockDateFilter(admin.SimpleListFilter):
    title = 'Date'
    parameter_name = 'date'

    def lookups(self, request, model_admin):
        dates = Stock.objects.datetimes('date_time', 'day', order='DESC')
        return [(date.strftime('%Y-%m-%d'), date.strftime('%Y-%m-%d')) for date in dates]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(date_time__date=self.value())
        return queryset
    
class StockAdmin(admin.ModelAdmin):
    list_display = ('symbol', 'date_time', 'open', 'high', 'low', 'close', 'previous_close')
    list_filter = (StockSymbolFilter, StockDateTimeFilter, StockDateFilter)

    def export_to_excel(self, request, queryset):
        queryset = queryset.order_by('date_time')  # Sort by datetime
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=stocks.xlsx'
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Symbol', 'Date Time', 'Open', 'High', 'Low', 'Close', 'Previous Close'])

        for obj in queryset:
            worksheet.append([obj.symbol, obj.date_time, obj.open, obj.high, obj.low, obj.close, obj.previous_close])

        workbook.save(response)
        return response

    export_to_excel.short_description = 'Export selected stocks to Excel'

    actions = ['export_to_excel']

admin.site.register(Stock, StockAdmin)